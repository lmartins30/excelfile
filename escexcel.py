import sys
import os
import pandas as pd
import openpyxl
from datetime import datetime

def excel():                                    #recebe o endereço do arquivo excel
 total = 0
 caminho_d = "Desktop\\down.xlsx"                # caminho do arquivo no pc
 planilha_d = "Plan1"
 caminho_u = "Desktop\\up.xlsx"                # caminho do arquivo no pc
 planilha_u = "Plan1"
 # Colunas do excel - Data, Nome, CPF, Login
 dado_d = pd.ExcelFile(caminho_d)
 ps_d = openpyxl.load_workbook(caminho_d)
 plan_d = ps_d[planilha_d]
 linhas = plan_d.max_row
 dado_u = pd.ExcelFile(caminho_u)
 ps_u = openpyxl.load_workbook(caminho_u)
 plan_u = ps_u[planilha_u]
 for row in range(1, linhas + 1):  # pega os dados das linhas do excel 
  # Pega os cabeçalhos da planilha
  if row == 1:                                      
   data = plan_d['A' + str(row)].value             # pega a coluna 1
   nome = plan_d['B' + str(row)].value             # pega a coluna 2
   idade  = plan_d['C' + str(row)].value           # pega a coluna 3
   login = plan_d['D' + str(row)].value            # pega a coluna 4
   modulo = plan_d['E' + str(row)].value
   modulo = checa_data(data)
   # Escreve no arquivo
   plan_u['A' + str(row)] = data
   plan_u['B' + str(row)] = nome
   plan_u['C' + str(row)] = cpf
   plan_u['D' + str(row)] = login
   plan_u['E' + str(row)] = modulo
   ps_u.save(caminho_u) 
  else:
   data = plan_d['A' + str(row)].value             # pega a data
   nome = plan_d['B' + str(row)].value             # pega o nome
   cpf  = plan_d['C' + str(row)].value             # pega o cpf
   login = plan_d['D' + str(row)].value            # pega o login
   modulo = plan_d['E' + str(row)].value
  # print("Aluno: ", end=' ')
  # print(row -1)
  # print(data)
  # print(nome)
  # print(cpf)
  # print(login)
  # print(modulo)
   data = data.date()
   modulo =  checa_data(data)
  
  # Escreve no arquivo de upload
   plan_u['A' + str(row)] = data
   plan_u['B' + str(row)] = nome
   plan_u['C' + str(row)] = cpf
   plan_u['D' + str(row)] = login
   plan_u['E' + str(row)] = modulo
   ps_u.save(caminho_u)
  

def checa_data(a):                           # Transforma a data + hora em somente data
 hoje = datetime.now().date()                # dia de consulta, no caso, hoje
 delta = (hoje - a).days                     # dias desde o cadastro
 if delta < 2:                               # se for menor que 2 dias
  # print ('liberar módulo 1')
  return  1
 elif delta >= 2 and delta < 5:              # se for mais ou exatamente 2 dias ou menos que 5 dias
  # print ('liberar módulo 2')
  return 2
 elif delta >= 5 and delta <= 7:             # se for mais ou exatamente 5 dias ou menos que 7 dias
  # print ('liberar módulo 3')
  return 3
 else:
  # print ('liberar todos módulos')
  return "ultimo numero de modulo do curso"

 
 
# This is the standard boilerplate that calls the main() function.
if __name__ == '__main__':
 excel()
 # excel("Desktop\\teste.xlsx")       # chamar a função excel com o endereço do arquivo excel